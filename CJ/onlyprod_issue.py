#외부모듈
import openpyxl
import os
from openpyxl import load_workbook
from openpyxl import Workbook

#내부모듈
from module import init


output_excel = Workbook()
output_sheet = output_excel.active
output_sheet.title = 'CJ'

input_excel = load_workbook(os.getcwd()+"/input/CJ_prodissue.xlsx", data_only=True)
input_sheet = input_excel['CJ']

init.sheet(input_sheet,output_sheet)#출력시트 초기화
cnt=2
while input_sheet.cell(cnt, 1).value:
    print(cnt)#행 번호출력
    raw = init.string(input_sheet,output_sheet,cnt)#원본행 로드
    raw = raw.replace("("," ").replace(")"," ")
    raw_clause = raw.split()
    prod = raw_clause[0]
    output_sheet.cell(cnt, 13, prod)  #생산자명
    cnt=cnt+1

output_excel.save(os.getcwd()+"/output/CJ_prodissue.xlsx")

