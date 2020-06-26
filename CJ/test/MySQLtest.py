import pymysql
from openpyxl import load_workbook
from openpyxl import Workbook
import os


input_excel = load_workbook(os.getcwd()+"/output/CJ.xlsx", data_only=True)
input_sheet = input_excel['CJ']

# MySQL Connection 연결
conn = pymysql.connect(host='localhost', user='root', password='123123',
                       db='mydb', charset='utf8')

# Connection 으로부터 Dictoionary Cursor 생성
curs = conn.cursor(pymysql.cursors.DictCursor)

# SQL문 실행
#sql = "create table msdb()"
#curs.execute(sql)

output_excel = Workbook()
output_sheet = output_excel.active
output_sheet.title = 'CJ'

os.chdir('..')
input_excel = load_workbook(os.getcwd() + "/output/CJ.xlsx", data_only=True)
input_sheet = input_excel['CJ']

cnt = 2
while input_sheet.cell(cnt, 1).value:
    sql = "insert msdb (1, %s %s %s %s %s"
    cnt = cnt + 1

# 데이타 Fetch
rows = curs.fetchall()
for row in rows:
    print(row)
    print(row['id'], row['name'])

# Connection 닫기
conn.close()