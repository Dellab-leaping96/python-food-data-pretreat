#외부모듈
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import pymysql.cursors
import os

#내부모듈
from CJ.test import create, select,insert

MSDB_excel = load_workbook(os.getcwd()+"/MSDB.xlsx", data_only=True)
MSDB_sheet = MSDB_excel['MSDB']
CJ_excel = load_workbook(os.getcwd()+"/CJ_1.xlsx", data_only=True)
CJ_sheet = CJ_excel['fail']
print("엑셀로드 완료")

conn = pymysql.connect(host='localhost',
                       user='root',
                       password='123123',
                       db='compdb',
                       charset='utf8mb4')
print("mysql 연결완료, 엔터누르면 진행")
input()

try:
    with conn.cursor() as cursor:
    
        #0. 테이블 초기화
        check = "SHOW TABLES LIKE 'msdb'"                                           # 기존에 msdb 테이블이 있으면 삭제
        cursor.execute(check)
        result=cursor.fetchall()
        if len(result) == 1:
            sql = "DROP TABLE msdb"
            cursor.execute(sql)
        print("테이블 초기화 완료\n")
        
        #1. 테이블 생성
        create.MSDB(cursor)                                                         # msdb 테이블 생성
        create.CJ(cursor)                                                           # CJ 테이블 생성
        print("테이블 생성 완료\n")

        #2. 테이블 값 입력
        insert.MSDB(cursor, MSDB_sheet)
        insert.CJ(cursor,CJ_sheet)

        print("테이블 입력 완료\n")

        #3. 테이블 조회
        print("========테이블 조회=======\n")
        #select.MSDB_ALL(cursor)
        select.CJ_ALL(cursor)
        print("========테이블 조회=======\n")

        #4. 테이블 삭제
        sql = 'DROP TABLE msdb'
        cursor.execute(sql)
        sql = 'DROP TABLE cj'
        cursor.execute(sql)
        print("테이블 삭제 완료\n")

        #5. DB 업데이트
        conn.commit()

finally:
    conn.close()
    print("연결 종료")