#외부모듈
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import re

#내부모듈
from CJ.module1 import init,pname,prod,stat,tax,wt,det
from CJ.module2 import keyprod, patternprod

output_excel = Workbook()
output_sheet = output_excel.active
output_sheet.title = 'CJ'

input_excel = load_workbook(os.getcwd()+"/output/CJ_1.xlsx", data_only=True)
input_sheet = input_excel['fail']

os.chdir('..')
producer_excel = load_workbook(os.getcwd()+"/prodDB.xlsx", data_only=True)
keyword_sheet = producer_excel['keyword']

init.sheet(input_sheet,output_sheet)#출력시트 초기화
cnt=2
while input_sheet.cell(cnt, 1).value:
    print(cnt)#행 번호출력
    raw = init.string(input_sheet,output_sheet,cnt)#원본행 로드

    productName, keyword_1 = pname.pre(raw)
    producer,keyword_2 = keyprod.pre(raw,keyword_sheet)
    status, keyword_3 = stat.pre(raw)
    taxataion, keyword_4 = tax.pre(raw)
    weight, keyword_5 = wt.pre(raw)
    detail = det.pre(raw, keyword_1, keyword_2, keyword_3, keyword_4, keyword_5)

    output_sheet.cell(cnt, 10, input_sheet.cell(cnt, 10).value)  #상품명
    output_sheet.cell(cnt, 11, input_sheet.cell(cnt, 11).value)  # 상품설명
    output_sheet.cell(cnt, 12, input_sheet.cell(cnt, 12).value)  # 중량
    output_sheet.cell(cnt, 13, producer)  #생산자명
    if producer == "":
        ptprodList = patternprod.match(raw, cnt, output_sheet)  # 생산자명 후보 작성
    output_sheet.cell(cnt, 14, input_sheet.cell(cnt, 14).value)  # 면과세여부
    output_sheet.cell(cnt, 15, input_sheet.cell(cnt, 15).value)  #보관상태


    cnt=cnt+1

os.chdir('CJ')
output_excel.save(os.getcwd()+"/output/CJ_2-1.xlsx")

