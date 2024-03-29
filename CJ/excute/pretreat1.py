#외부모듈
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import re

#내부모듈
from CJ.module1 import init,pname,prod,stat,tax,wt,det

output_excel = Workbook()
output_sheet = output_excel.active
output_sheet.title = 'CJ'

os.chdir('..')
input_excel = load_workbook(os.getcwd()+"/input/CJ.xlsx", data_only=True)
input_sheet = input_excel['CJ']

os.chdir('..')
producer_excel = load_workbook(os.getcwd()+"/ProdDB.xlsx", data_only=True)
maker_sheet = producer_excel['maker']
nation_sheet = producer_excel['nation']

init.sheet(input_sheet,output_sheet)#출력시트 초기화
cnt=2
while input_sheet.cell(cnt, 1).value:
    print(cnt)#행 번호출력
    raw = init.string(input_sheet,output_sheet,cnt)#원본행 로드

    productName,keyword_1 = pname.pre(raw)
    producer,keyword_2 = prod.pre(raw,maker_sheet, nation_sheet)
    status,keyword_3 = stat.pre(raw)
    taxataion,keyword_4 = tax.pre(raw)
    weight,keyword_5= wt.pre(raw)
    detail = det.pre(raw, keyword_1,keyword_2,keyword_3,keyword_4,keyword_5)

    output_sheet.cell(cnt, 10, productName)  #상품명
    output_sheet.cell(cnt, 11, detail)  # 상품설명
    output_sheet.cell(cnt, 12, weight)  # 중량
    output_sheet.cell(cnt, 13, producer)  #생산자명
    output_sheet.cell(cnt, 14, taxataion)  # 면과세여부
    output_sheet.cell(cnt, 15, status)  #보관상태


    cnt=cnt+1

os.chdir('CJ')
output_excel.save(os.getcwd()+"/output/CJ_1.xlsx")

