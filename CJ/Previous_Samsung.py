#한글깨짐 해결
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

#엑셀편집모듈
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os

#자체제작모듈
from 내모듈 import 중량치환
from 내모듈 import 시트초기화
from 내모듈 import 면과세_보관
from 내모듈 import 생산자명추출
from 내모듈 import 특수문자제거


#출력파일 만들기
write_wb = Workbook()
ws = write_wb.active
ws.title = 'sample'
write_wb.remove(write_wb['sample'])
write_ws = write_wb.create_sheet("삼성")#시트생성

#원본파일 로드
load_wb = load_workbook(os.getcwd()+"/원본파일/삼성원본.xlsx", data_only=True)
load_ws = load_wb["삼성"]
maker_wb = load_workbook(os.getcwd()+"/원본파일/MSDB 정의서.xlsx")
maker_ws = maker_wb['제조사명DB']
nation_ws = maker_wb['국가명DB']
시트초기화.init(load_ws, write_ws)

#1행씩 읽으면서 출력
num = 2
while True:
    print("=====", num, "=====")

    #입력값 로드
    String = ""  # 모든 열의 정보가 합쳐진 문자열로 만들기
    for i in range(1, 5):
        if load_ws.cell(num, i).value:
            String += str(load_ws.cell(num, i).value) + "/"

    #잘못된 정보들 치환
    String = String.replace("㎏", "kg").replace("㎖", "ml").replace("ℓ", "l")

    #String = String.replace("/"," ")

    #상품명추출
    name =

    #면과세추출
    tax = 면과세추출.Tax(String)

    #보관상태추출
    stat = 보관상태추출.Stat(String)

    #생산자명추출
    maker = 생산자명추출.Maker(String,tax,maker_ws, nation_ws)

    #중량추출
    weight = 중량치환.Weight(String)

    """
    ###각 열 별로 분리,추출

    # 상세정보분리
    String, detail = Find_Detail(String)
    if maker_2 != "" and detail != "":
        detail = detail + "," + maker_2
    elif maker_2 != "" and detail == "":
        detail = maker_2

    # 최종정보에서 필요없는값 삭제
    String = String.replace("ea", "").replace("kg", "")

    # 생산자명 없을경우 상품명에서 한번더 분리
    if maker_1 == "":
        String, maker_1 = Find_Maker_all(String, maker_ws)

    """

    ###원본데이터 작성
    for i in range(1, 9):
        if load_ws.cell(num, i).value:
            write_ws.cell(num, i, str(load_ws.cell(num, i).value))

    ###분리한 정보들 작성
    #write_ws.cell(num, 10, String)  # 상품명
    #write_ws.cell(num, 11, detail)  # 상품상세
    write_ws.cell(num, 12, weight)  # 상품중량
    write_ws.cell(num, 13, maker)  # 생산자명
    write_ws.cell(num, 14, tax)  # 면세여부
    write_ws.cell(num, 15, stat)  # 보관상태

    num = num + 1

    # 인풋값 다 읽으면 반복문 종료
    if not load_ws.cell(num, 1).value:
        break

# 전처리 완료값 저장
write_wb.save(os.getcwd()+"/출력파일/삼성.xlsx")
