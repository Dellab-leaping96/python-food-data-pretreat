#외부모듈
import openpyxl
import os
from openpyxl import load_workbook
from openpyxl import Workbook

#내부모듈
from module import init
from module import prod


output_excel = Workbook()
output_sheet = output_excel.active
output_sheet.title = '삼성'

input_excel = load_workbook(os.getcwd()+"/input/CJ.xlsx", data_only=True)
input_sheet = input_excel['CJ']

os.chdir('..')
producer_excel = load_workbook(os.getcwd()+"/ProdDB.xlsx", data_only=True)
maker_sheet = producer_excel['maker']
nation_sheet = producer_excel['nation']

init.sheet(input_sheet,output_sheet)#출력시트 초기화
cnt=2
while input_sheet.cell(cnt, 1).value:
    print(cnt)#행 번호출력
    raw = init.string(input_sheet,output_sheet,cnt)#원본행 로드
    producer = prod.pre(raw,maker_sheet, nation_sheet)
    output_sheet.cell(cnt, 13, producer)  #생산자명
    cnt=cnt+1

os.chdir('CJ')
output_excel.save(os.getcwd()+"/output/CJ_prod.xlsx")

