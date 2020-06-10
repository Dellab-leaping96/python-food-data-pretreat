import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import re
from openpyxl.styles import PatternFill


# 특수문자를 모두 공백으로 바꾸는 함수
def cleanText(readData):
    text = re.sub(
        '[-=+,#/\?:^$.@*\"※~%ㆍ!』\\‘|\(\)\[\]\<\>`\'…》_]', ' ', readData)
    return text

# 국가명 찾고 분리하는 함수
def Find_Nation(String, nation_ws):
    String_touse = cleanText(String)
    String_iter = String_touse.split()
    maker_num = 2
    maker = ""
    while True:
        if nation_ws.cell(maker_num, 1).value is None:
            break
        for element in String_iter:
            if element == str(nation_ws.cell(maker_num, 1).value):
                String = String.replace(element, "")
                maker = str(nation_ws.cell(maker_num, 2).value)
                break
        maker_num += 1
    return String, maker


# 제조사명 찾고 분리하는 함수
# 어절단위로만 찾음
def Find_Maker_clause(String, maker_ws):
    String_touse = cleanText(String)
    String_iter = String_touse.split()
    maker_num = 2
    maker = ""
    while True:
        if maker_ws.cell(maker_num, 1).value is None:
            break
        for element in String_iter:  # 규격탭에 없을경우 상품명탭에서 한번더 국가찾기
            if element == str(maker_ws.cell(maker_num, 1).value):
                String = String.replace(element, "")
                maker = str(maker_ws.cell(maker_num, 2).value)
                # break
                return String, maker
        maker_num += 1

    return String, maker


# 그냥 문자전체에서 제조사명 있기만하면 분리하는함수
def Find_Maker_all(String, maker_ws):
    # String_touse = cleanText(String)
    maker_num = 2
    maker = ""
    while True:
        if maker_ws.cell(maker_num, 1).value is None:
            break

        if String.find(str(maker_ws.cell(maker_num, 1).value)) != -1:
            String = String.replace(str(maker_ws.cell(maker_num, 1).value), "")
            maker = str(maker_ws.cell(maker_num, 2).value)
            # break
            return String, maker
        maker_num += 1

    return String, maker


# 면세여부 찾고 분리하는 함수
def FInd_Tax(String):
    String_iter = String.split()
    tax = ""
    for element in String_iter:
        if element == "과세":
            String = String.replace(element, "")
            tax = "N"
        elif element == "면세":
            String = String.replace(element, "")
            tax = "Y"
    return String, tax


# 보관상태 찾고 분리하는 함수
def Find_stat(String):
    String_copy = cleanText(String)
    String_iter = String_copy.split()
    stat = ""
    for element in String_iter:
        if element == ("실온"):
            String = String.replace(element, "")
            stat = "상온"
        elif element == ("냉장"):
            String = String.replace(element, "")
            stat = "냉장"
        elif element == ("냉동"):
            String = String.replace(element, "")
            stat = "냉동"

    if stat == "":
        if String.find("상온")!=-1:
            String.replace("상온","")
            stat = "상온"
        if String.find("실온")!=-1:
            String.replace("실온","")
            stat = "상온"
        if String.find("냉장")!=-1:
            String.replace("냉장","")
            stat = "냉장"
        if String.find("냉동")!=-1:
            String.replace("냉동","")
            stat = "냉동"
    String = String.replace("상온", "").replace("냉장", "").replace("냉동", "")


    return String, stat


# 중량 찾고 분리하는 함수
def Find_weight(String):
    weight = ""

    if String == "":
        print("오류")

    else:
        String = String.lower()
        String_touse = String.lower()  # 모두 소문자화
        # 특수문자 제거
        String_touse = String_touse.replace("(", " ")
        String_touse = String_touse.replace(")", " ")
        String_touse = String_touse.replace("_", " ")
        String_touse = String_touse.replace("/", " ")
        String_touse = String_touse.replace(",", "")

        String_iter = String_touse.split()
        String_touse = ""
        for element in String_iter:
            if element.find("*") == -1 and element.find("~") and element.find("∼") == -1:
                String_touse += element + " "
        pattern = r"(([0-9]+)?(\.)?[0-9]+(m?l|k?g))"
        i = re.finditer(pattern, String_touse)  # 패턴매칭, 리스트로작성
        weight1 = []
        weight2 = []

        for element in i:
            weight += " " + element.group()

        weight1 = weight.split()
        weight_num = 0

        for element in weight1:
            is_k = 0
            element = element.replace("g", "")
            element = element.replace("ml", "")

            if element.find("k") != -1 or element.find("l") != -1:
                is_k = 1
                element = element.replace("k", "")
                element = element.replace("l", "")

            element = float(element)
            if is_k == 1:
                element = int(element * 1000)
            else:
                element = int(element)
            weight2.append(element)
            weight_num = weight_num + 1

            weight_num = 0

            index = 0
            for element in weight2:
                if element == max(weight2):
                    index = weight_num
                weight_num = weight_num + 1

            if weight1:
                weight = weight1[index]

        if weight == "" and String_touse.find("kg") != -1:
            String = String.replace("kg", "")
            weight = "1kg"
        else:
            String = String.replace(weight, "")

    if weight == "" and String.find("kg") != -1:
        String = String.replace("1kg", "").replace("1KG", "").replace("kg", "".replace("KG", ""))
        weight = "1kg"
        String = String.lower()

    return String, weight


# 상세정보 찾고 분리하는 함수
def Find_Detail(String):
    all_details = []  # 모든 디테일들 저장할 배열
    out_detail = ""  # 최종 상품상세
    name = ""  # 최종 상품명
    String = String.replace("_", " ").replace("/", " ")

    # ()괄호로 묶인 문장들 모두 분리하여 배열에 옮김
    String = String.replace("()", " ")
    # extracts string in bracket()
    items1 = re.finditer('\(([^)]+)\)', String)
    for element in items1:
        String = String.replace(element.group(), " ")
        out_element = (element.group()).replace(
            "(", "").replace(")", "").replace(",", " ").replace(":", "")
        out_elements = out_element.split()
        for e in out_elements:
            all_details.append(e)

    # []괄호로 묶인 문장들 모두 분리하여 배열에 옮김
    String = String.replace("[]", " ")
    items2 = re.finditer('\[([^)]+)\]', String)
    for element in items2:
        String = String.replace(element.group(), " ")
        out_element = (element.group()).replace(
            "[", "").replace("]", "").replace(",", " ").replace(":", "")
        out_elements = out_element.split()
        for e in out_elements:
            all_details.append(e)

    # 남은 스트링에서 첫번째 어절만 상품명으로 취하고 나머지정보는 정리해서 상세정보로 처리
    devide = String.split()
    if devide:
        String = devide[0]  # 첫번째 어절 상품명
        for i in range(1, len(devide)):  # 나머지어절은 상세정보 리스트에 추가
            all_details.append(devide[i].replace(
                "/", ",").replace(",", " ").replace(" ", ",").replace(":", ""))

    # 모든 상세정보리스트 종합
    for i in range(0, len(all_details)):
        all_details[i] = all_details[i].replace(",", "").replace("(", "").replace(")", "").replace("[", ""
                                                                                                   ).replace("]", "")

        # 단순 정보는 제거
        if all_details[i] == "ea" or all_details[i] == "팩" or all_details[i
        ] == "곽" or all_details[i] == "pet" or all_details[i] == "두께" or all_details[i
        ] == "/" or all_details[i] == "-" or all_details[i] == "기타" or all_details[
            i] == "통" or all_details[i] == "pac" or all_details[i] == "pak" or all_details[
            i] == "kg" or all_details[i] == "-n" or all_details[i] == "개" or all_details[i
        ] == "봉" or all_details[i] == "팩포장" or all_details[i] == "손":
            all_details[i] = ""

    all_details = list(set(all_details))
    for i in range(0, len(all_details)):
        if i == len(all_details) - 1:
            out_detail += all_details[i]
        else:
            out_detail += all_details[i] + ","
    out_detail = out_detail.replace(",,", ",").replace(",,", ","
                                                       ).replace(",,", ",").replace(",,", ",").replace(",,", ",")
    out_detail = out_detail.strip(',')

    return String, out_detail


# 전처리하는 전체함수
def pretreat(excelName, company):
    # 엑셀 출력물 초기세팅
    write_wb = Workbook()
    ws = write_wb.active
    ws.title = 'sample'
    write_wb.remove(write_wb['sample'])
    write_ws = write_wb.create_sheet(company)
    init_sheet(write_ws)

    # 엑셀 입력물 로드
    load_wb = load_workbook(excelName, data_only=True)
    load_ws = load_wb[company]

    maker_wb = load_workbook('C:\\Users\\딜리버리랩\\Desktop\\MSDB 정의서.xlsx')
    maker_ws = maker_wb['제조사명DB']
    nation_ws = maker_wb['국가명DB']

    # 한줄씩 처리
    num = 2

    while True:
        print("=====", num, "=====")

        ###엑셀 입력값 가져옴
        String = ""  # 모든 열의 정보가 합쳐진 문자열로 만들기
        for i in range(1, 5):
            if load_ws.cell(num, i).value:
                String += str(load_ws.cell(num, i).value) + " "

        ###원본데이터 열(탭)들 명시
        for i in range(1, 9):
            if load_ws.cell(1, i).value:
                # 전처리에 사용된 열(탭)들 명시 + 단가&상품코드명시
                # 1~5열은 전처리에 사용될 문자열 / 6열은 단가 / 7열은 상품코드 / 8열은 분류로 약속해뒀음
                write_ws.cell(1, i, str(load_ws.cell(1, i).value))

        ###원본데이터 작성(비교용)
        for i in range(1, 9):
            if load_ws.cell(num, i).value:
                write_ws.cell(num, i, str(load_ws.cell(num, i).value))

        ###출력할 값들 변수선언
        name = ""  # 상품명
        detail = ""  # 상품상세
        weight = ""  # 중량
        maker_1 = ""  # 생산자명1
        maker_2 = ""  # 생산자명2
        tax = ""  # 면세여부
        stat = ""  # 보관상태

        ###필요없는 정보들 제거
        if company == "현대":
            # 현대기준제거
            String = String.replace("/BOX", " ").replace("/EA", " ").replace("/PK", " ").replace("/Box", " ")
            String = String.replace("/box", " ").replace("/ea", " ").replace("/pk", " ")
            String = String.replace("/포", " ").replace("/통", " ").replace("/봉", " ").replace("/장", " ")

        elif company == "CJ":
            # CJ기준제거
            String = String.replace("/EA", "").replace("/개", "").replace("/PAC", "").replace("/PAK", "").replace("/BOX",
                                                                                                                 "")

        elif company == "아워홈":
            # 아워홈기준제거
            String = String.replace("/BOX", "").replace("BOX", "").replace("box", "").replace("Box", ""
                                                                                              ).replace("PK.",
                                                                                                        "").replace(
                "pk.", "").replace("-", " ").replace("/봉", "")

        elif company == "동원":
            # 동원기준제거
            String = String.replace("box", "").replace("pk", "").replace("/", " ").replace("BOX", " ").replace(
                ")PK", ")").replace(")PAK", ")").replace(")EA", ")")

        elif company == "한화":
            # 한화기준제거
            String = String.replace("국내/", "").replace("수입/", ""
                                                       ).replace("BOX/", "").replace("pk", "").replace("EA/", ""
                                                                                                       ).replace("BOX",
                                                                                                                 "").replace(
                "EA/", "").replace("/", " "
                                   )

        ###잘못된 정보들 변환
        String = String.replace("㎏", "kg").replace("㎖", "ml").replace("ℓ", "l")

        ###각 열 별로 분리,추출

        # 면세여부 분리
        String, tax = FInd_Tax(String)

        # 보관상태 분리
        String, stat = Find_stat(String)

        # 생산자명분리
        if tax == "N":  # 과세일경우 제조사명,국가명 순으로 분리
            String, maker_1 = Find_Maker_clause(String, maker_ws)
            String, maker_2 = Find_Nation(String, nation_ws)
        elif tax == "Y":  # 면세일경우 국가명,제조사명 순으로 분리
            String, maker_1 = Find_Nation(String, nation_ws)
            String, maker_2 = Find_Maker_clause(String, maker_ws)

        if maker_1 == "": #생산자명 추출 안돨경우 상세로 갈 값을 매칭
            maker_1 = maker_2
            maker_2 = ""

        # 총중량 분리
        String, weight = Find_weight(String)

        # 상세정보분리
        String, detail = Find_Detail(String)
        if maker_2 != "" and detail != "":
            detail = detail + "," + maker_2
        elif maker_2 != "" and detail == "":
            detail = maker_2

        # 최종정보에서 필요없는값 삭제
        String = String.replace("ea", "").replace("kg", "")

        # 생산자명 없을경우 상품명에서 한번더 분리
        if maker_1 == "":
            String, maker_1 = Find_Maker_all(String, maker_ws)

        # print(maker_1, "||", maker_2, "||", tax, "||", stat,"||", weight, "||", detail, "||", String)
        # input()

        ###분리한 정보들 작성
        write_ws.cell(num, 10, String)  # 상품명
        write_ws.cell(num, 11, detail)  # 상품상세
        write_ws.cell(num, 12, weight)  # 상품중량
        write_ws.cell(num, 13, maker_1)  # 생산자명
        write_ws.cell(num, 14, tax)  # 면세여부
        write_ws.cell(num, 15, stat)  # 보관상태

        num = num + 1

        # 인풋값 다 읽으면 반복문 종료
        if not load_ws.cell(num, 1).value:
            break

    # 전처리 완료값 저장
    write_wb.save("C:\\Users\\딜리버리랩\\Desktop\\자동분류\\13주차\\" + company + "_output.xlsx")
    return


###메인###
pretreat("C:\\Users\\딜리버리랩\\Desktop\\자동분류\\13주차\\1차전처리\\MSDB작업용(유통코드)_v2.2.4.xlsx", "삼성")
pretreat("C:\\Users\\딜리버리랩\\Desktop\\자동분류\\13주차\\1차전처리\\MSDB작업용(유통코드)_v2.2.4.xlsx", "CJ")
pretreat("C:\\Users\\딜리버리랩\\Desktop\\자동분류\\13주차\\1차전처리\\MSDB작업용(유통코드)_v2.2.4.xlsx", "한화")
pretreat("C:\\Users\\딜리버리랩\\Desktop\\자동분류\\13주차\\1차전처리\\MSDB작업용(유통코드)_v2.2.4.xlsx", "현대")
pretreat("C:\\Users\\딜리버리랩\\Desktop\\자동분류\\13주차\\1차전처리\\MSDB작업용(유통코드)_v2.2.4.xlsx", "동원")
pretreat("C:\\Users\\딜리버리랩\\Desktop\\자동분류\\13주차\\1차전처리\\MSDB작업용(유통코드)_v2.2.4.xlsx", "아워홈")
